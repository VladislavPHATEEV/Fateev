import csv
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

#ЧТО ТО ОДНО

class Vacancy:
    curr_to_rub = {
		"KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055,
        "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
    }
    def __init__(self, vacancy):
        self.salary_to = int(float(vacancy['salary_to']))
		self.salary_from = int(float(vacancy['salary_from']))
        self.salary_average = self.curr_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
		self.salary_currency = vacancy['salary_currency']
        self.year = int(vacancy['published_at'][:4])
		self.name = vacancy['name']
		self.area_name = vacancy['area_name']
class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')
        dataset = DataSet(self.file_name, self.vacancy_name)
        stats1, stats2, stats3, stats4, stats5, stats6 = dataset.get_stat()
        dataset.print_stats(stats1, stats2, stats3, stats4, stats5, stats6)
        report = Report(self.vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6)
        report.gen_excel()
class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.vacancy_name = vacancy_name
		        self.file_name = file_name
	@staticmethod
    def avrg(dictionary):
        new_dictionary = {}
        for key, values in dictionary.items():
            new_dictionary[key] = int(sum(values) / len(values))
        return new_dictionary			
    @staticmethod
    def incr(dictionary, key, amount):
        if key not in dictionary: dictionary[key] = amount
        else: dictionary[key] += amount
	@staticmethod
    def print_stats(stats1, stats2, stats3, stats4, stats5, stats6):
        print('Динамика уровня зарплат по годам: {0}'.format(stats1))
        print('Динамика количества вакансий по годам: {0}'.format(stats2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(stats3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(stats4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(stats5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(stats6))
    def csv_reader(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
			len_head = len(header)
            header_length = len_head
            for row in reader:
                if '' not in row and len(row) == header_length:
                    yield dict(zip(header, row))
    def get_stat(self):
		salary_city = {}
		salary_of_vacancy_name = {}
        salary = {}
        count_of_vacancies = 0
        for vacancy_dictionary in self.csv_reader():
            vacancy = Vacancy(vacancy_dictionary)
            self.incr(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.incr(salary_of_vacancy_name, vacancy.year, [vacancy.salary_average])
            self.incr(salary_city, vacancy.area_name, [vacancy.salary_average])
            count_of_vacancies += 1
        vacancies_number = dict([(key, len(value)) for key, value in salary.items()])
        vacancies_number_by_name = dict([(key, len(value)) for key, value in salary_of_vacancy_name.items()])
        if not salary_of_vacancy_name:
            salary_of_vacancy_name = dict([(key, [0]) for key, value in salary.items()])
            vacancies_number_by_name = dict([(key, 0) for key, value in vacancies_number.items()])
		stats2 = self.avrg(salary_of_vacancy_name)
        stats3 = self.avrg(salary_city)
        stats = self.avrg(salary)
        stats4 = {}
        for year, salaries in salary_city.items():
            stats4[year] = round(len(salaries) / count_of_vacancies, 4)
        stats4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in stats4.items()]))
        stats4.sort(key=lambda a: a[-1], reverse=True)
        stats5 = stats4.copy()
		aab = dict(stats4)
        stats4 = aab
        stats3 = list(filter(lambda a: a[0] in list(stats4.keys()), [(key, value) for key, value in stats3.items()]))
        stats3.sort(key=lambda a: a[-1], reverse=True)
        stats3 = dict(stats3[:10])
        stats5 = dict(stats5[:10])
        return stats, vacancies_number, stats2, vacancies_number_by_name, stats3, stats5
class Report:
    def __init__(self, vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6):
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
		self.stats2 = stats2
        self.stats3 = stats3
        self.stats1 = stats1
		self.stats5 = stats5
        self.stats6 = stats6
        self.stats4 = stats4
    def gen_excel(self):
        wwwe1 = self.wb.active
        wwwe1.title = 'Статистика по годам'
        wwwe1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for year in self.stats1.keys():
            wwwe1.append([year, self.stats1[year], self.stats3[year], self.stats2[year], self.stats4[year]])
        dataa = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in dataa:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]
        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            wwwe1.column_dimensions[get_column_letter(i)].width = column_width + 2
        dataa = []
        dataa.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.stats5.items(), self.stats6.items()):
            dataa.append([city1, value1, '', city2, value2])
        wwwwel1 = self.wb.create_sheet('Статистика по городам')
        for row in dataa:
            wwwwel1.append(row)
        column_widths = []
        for row in dataa:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]
        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            wwwwel1.column_dimensions[get_column_letter(i)].width = column_width + 2
        font_bold = Font(bold=True)
        for col in 'ABCDE':
            wwwe1[col + '1'].font = font_bold
            wwwwel1[col + '1'].font = font_bold
        for index, _ in enumerate(self.stats5):
            wwwwel1['E' + str(index + 2)].number_format = '0.00%'
        thin = Side(border_style='thin', color='00000000')
        for row in range(len(dataa)):
            for col in 'ABDE':
                wwwwel1[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)
        self.stats1[1] = 1
        for row, _ in enumerate(self.stats1):
            for col in 'ABCDE':
                wwwe1[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)
        self.wb.save('report.xlsx')
if __name__ == '__main__':
    InputConnect()